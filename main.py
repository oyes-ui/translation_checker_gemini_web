from dotenv import load_dotenv
# Load environment variables FIRST before other imports
load_dotenv()

from fastapi import FastAPI, UploadFile, File, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import shutil
import os
import uuid
import asyncio
import json
import openpyxl
from contextlib import asynccontextmanager
from checker_service import TranslationChecker

# Directory for temp files
UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

app = FastAPI()

# 1. 현재 main.py 파일이 있는 실제 경로를 계산합니다.
current_dir = os.path.dirname(os.path.abspath(__file__))
static_dir = os.path.join(current_dir, "static")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



# Store for active tasks: { task_id: { "queue": asyncio.Queue, "result_path": str, "status": ... } }
TASK_STORE = {}

class StartRequest(BaseModel):
    source_file_id: str
    target_file_id: str
    glossary_file_id: str = None
    sheets: list[str] = None
    sheet_langs: dict = {} # {"Sheet1": {"lang": "Korean", "code": "ko_KR"}}
    glossary_url: str = "https://docs.google.com/spreadsheets/d/1kVEdSTqZcFHLK8tK6IsF3Jb5ks-42RQgDimZ-rziKxU/gviz/tq?tqx=out:csv&sheet=용어집%20DB"
    source_lang: str = "English"
    target_lang: str = "Korean"
    target_code: str = "ko_KR"
    max_concurrency: int = 5
    cell_range: str = "C7:C28" # Default range
    model_name: str

async def background_inspection_task(task_id, params):
    queue = TASK_STORE[task_id]["queue"]
    try:
        checker = TranslationChecker(
            model_name=params.model_name,
            max_concurrency=params.max_concurrency,
            skip_llm_when_glossary_mismatch=True,
            no_backtranslation=True # Disabled as per user request
        )
        
        source_path = os.path.join(UPLOAD_DIR, params.source_file_id)
        target_path = os.path.join(UPLOAD_DIR, params.target_file_id)
        
        glossary_path = None
        if params.glossary_file_id:
            glossary_path = os.path.join(UPLOAD_DIR, params.glossary_file_id)

        # Run generator
        gen = checker.run_inspection_async_generator(
            source_file_path=source_path,
            target_file_path=target_path,
            cell_range=params.cell_range,
            source_lang=params.source_lang,
            target_lang=params.target_lang,
            target_lang_code=params.target_code,
            sheet_lang_map=params.sheet_langs,
            glossary_url=params.glossary_url,
            glossary_file_path=glossary_path,
            selected_sheets=params.sheets
        )
        
        async for event in gen:
            # If complete, save output
            if event["type"] == "complete":
                output_filename = f"translation_review_{task_id}.txt"
                output_path = os.path.join(UPLOAD_DIR, output_filename)
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(event["output_data"])
                
                TASK_STORE[task_id]["result_path"] = output_path
                # Notify completion without large data payload
                await queue.put({"type": "complete", "download_url": f"/api/download/{task_id}"})
            else:
                await queue.put(event)
                
    except Exception as e:
        print(f"Background Task Error: {e}")
        await queue.put({"type": "error", "message": f"검수 시작 중 오류 발생: {str(e)}"})
    finally:
        # cleanup files maybe? or keep them for now.
        pass

@app.post("/api/upload")
async def upload_files(
    source: UploadFile = File(None),
    target: UploadFile = File(None),
    glossary: UploadFile = File(None)
):
    result = {}
    
    try:
        # Handle Source Upload
        if source:
            s_id = f"src_{uuid.uuid4()}.xlsx"
            s_path = os.path.join(UPLOAD_DIR, s_id)
            content = await source.read()
            if not content:
                raise Exception("Uploaded source file is empty.")
            
            with open(s_path, "wb") as f:
                f.write(content)
                
            # Extract sheets from Source
            try:
                print(f"Opening workbook at {s_path}...")
                # data_only=True is usually what we want for checking values
                wb = openpyxl.load_workbook(s_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                print(f"Sheets found: {sheets}")
                wb.close()
            except Exception as e:
                print(f"Error loading workbook: {e}")
                # Re-try without read_only as some files might need it
                try:
                    wb = openpyxl.load_workbook(s_path, data_only=True)
                    sheets = wb.sheetnames
                    wb.close()
                except Exception as e2:
                    raise Exception(f"Failed to read Excel sheets: {str(e2)}")
                
            result["source_file_id"] = s_id
            result["sheets"] = sheets

        # Handle Target Upload
        if target:
            t_id = f"tgt_{uuid.uuid4()}.xlsx"
            t_path = os.path.join(UPLOAD_DIR, t_id)
            content = await target.read()
            if not content:
                 raise Exception("Uploaded target file is empty.")
            with open(t_path, "wb") as f:
                f.write(content)
            result["target_file_id"] = t_id

        # Handle Glossary Upload
        if glossary:
            g_id = f"glossary_{uuid.uuid4()}.csv"
            g_path = os.path.join(UPLOAD_DIR, g_id)
            content = await glossary.read()
            if not content:
                raise Exception("Uploaded glossary file is empty.")
            with open(g_path, "wb") as f:
                f.write(content)
            result["glossary_file_id"] = g_id

        if not result:
             raise HTTPException(status_code=400, detail="No files provided in the request.")
             
        return result

    except Exception as e:
        print(f"Upload error: {e}")
        raise HTTPException(status_code=400, detail=str(e))


class GlossaryCheckRequest(BaseModel):
    url: str
    source_lang: str = "English"

@app.post("/api/check_glossary")
async def check_glossary(req: GlossaryCheckRequest):
    checker = TranslationChecker(max_concurrency=1)
    msg = await checker.load_glossary_from_url(req.url, req.source_lang)
    # msg format is "Glossary loaded: N entries." or "Glossary load failed: ..."
    if "failed" in msg.lower():
        raise HTTPException(status_code=400, detail=msg)
    return {"message": msg}

@app.post("/api/start")
async def start_inspection(req: StartRequest, background_tasks: BackgroundTasks):
    task_id = str(uuid.uuid4())
    TASK_STORE[task_id] = {
        "queue": asyncio.Queue(),
        "result_path": None
    }
    
    background_tasks.add_task(background_inspection_task, task_id, req)
    
    return {"task_id": task_id}

@app.get("/api/stream/{task_id}")
async def stream_progress(task_id: str):
    if task_id not in TASK_STORE:
        raise HTTPException(status_code=404, detail="Task not found")
        
    async def event_generator():
        queue = TASK_STORE[task_id]["queue"]
        while True:
            data = await queue.get()
            yield f"data: {json.dumps(data)}\n\n"
            if data["type"] in ["complete", "error"]:
                break
    
    return StreamingResponse(event_generator(), media_type="text/event-stream")

@app.get("/api/download/{task_id}")
async def download_result(task_id: str):
    if task_id not in TASK_STORE or not TASK_STORE[task_id]["result_path"]:
        raise HTTPException(status_code=404, detail="Result not ready or task not found")
    
    path = TASK_STORE[task_id]["result_path"]
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Result file missing from server")
    
    return FileResponse(
        path, 
        filename=f"translation_review_{task_id}.txt", 
        media_type="application/octet-stream"
    )

# 1. 'static' 폴더를 웹에 연결
app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")

# 2. 접속 시 첫 화면(index.html) 보내주기
@app.get("/")
async def read_index():
    return FileResponse('static/index.html')

# 3. 브라우저 favicon.ico 404 에러 방지
@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    from fastapi.responses import Response
    return Response(status_code=204)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
